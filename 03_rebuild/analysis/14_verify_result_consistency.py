#!/usr/bin/env python
"""Verify consistency between manuscript claims, result tables and package files."""

from __future__ import annotations

import csv
import hashlib
import math
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
REBUILD_DIR = SCRIPT_PATH.parents[1]
ROOT = REBUILD_DIR.parent
RESULTS = REBUILD_DIR / "results"
VALIDATION = REBUILD_DIR / "validation"
MANUSCRIPT_DIR = REBUILD_DIR / "manuscript"
REFINED_MANUSCRIPT = MANUSCRIPT_DIR / "SCI_Q1_REFINED_MANUSCRIPT.md"
REFINED_MANUSCRIPT_FALLBACK = MANUSCRIPT_DIR / "SCI_Q1_REFINED_MANUSCRIPT.md"
REFINED_MANUSCRIPT = MANUSCRIPT_DIR / "SCI_Q1_REFINED_MANUSCRIPT.md"
BASE_MANUSCRIPT = MANUSCRIPT_DIR / "MANUSCRIPT_MAIN_TEXT.md"
MANUSCRIPT = Path(os.environ["GSE200996_MANUSCRIPT_PATH"]).resolve() if os.environ.get("GSE200996_MANUSCRIPT_PATH") else (
    REFINED_MANUSCRIPT
    if REFINED_MANUSCRIPT.exists()
    else (
        REFINED_MANUSCRIPT_FALLBACK
        if REFINED_MANUSCRIPT_FALLBACK.exists()
        else (REFINED_MANUSCRIPT if REFINED_MANUSCRIPT.exists() else BASE_MANUSCRIPT)
    )
)
FIGURE_SOURCE = REBUILD_DIR / "figures" / "submission" / "source_data"
TABLE_DIR = REBUILD_DIR / "tables" / "submission"
PACKAGE_DIR = REBUILD_DIR / "submission_package"
ZIP_PATH = PACKAGE_DIR / "OSCC_GSE200996_submission_bundle.zip"
OUT_MD = MANUSCRIPT_DIR / "RESULT_CONSISTENCY_AUDIT.md"
OUT_CSV = MANUSCRIPT_DIR / "RESULT_CONSISTENCY_AUDIT.csv"


@dataclass
class Check:
    area: str
    item: str
    status: str
    detail: str


checks: list[Check] = []


def add(area: str, item: str, status: str, detail: str):
    checks.append(Check(area, item, status, detail))


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def fnum(value: str) -> float:
    if value is None or str(value).strip().upper() in {"", "NA", "NAN", "NULL"}:
        return math.nan
    return float(value)


def finite_less(value: str, threshold: float) -> bool:
    parsed = fnum(value)
    return math.isfinite(parsed) and parsed < threshold


def close(actual: float, expected: float, rel_tol=5e-3, abs_tol=5e-4) -> bool:
    return math.isclose(actual, expected, rel_tol=rel_tol, abs_tol=abs_tol)


def row_by(rows: list[dict[str, str]], column: str, value: str) -> dict[str, str]:
    for row in rows:
        if row.get(column) == value:
            return row
    raise KeyError(f"{value} not found in {column}")


def check_numeric(area: str, item: str, actual: float, expected: float, source: Path, tol=5e-3):
    status = "PASS" if close(actual, expected, rel_tol=tol) else "FAIL"
    add(area, item, status, f"actual={actual:.8g}; expected={expected:.8g}; source={rel(source)}")


def check_text_contains(text: str, label: str, patterns: list[str]):
    missing = [pattern for pattern in patterns if pattern not in text]
    add(
        "Manuscript text",
        label,
        "PASS" if not missing else "FAIL",
        "All expected text fragments present." if not missing else "Missing: " + "; ".join(missing),
    )


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def audit_data_boundary(text: str):
    summary = (RESULTS / "data_audit" / "DATA_AUDIT_SUMMARY.md").read_text(encoding="utf-8")
    expected_fragments = {
        "Raw h5 files retained": "41",
        "Tumor h5 files": "25",
        "PBMC h5 files": "16",
        "CD45 tumor immune cells in metadata": "74557",
        "Patients in CD45 tumor metadata": "19",
        "Strict pre/post paired patients in CD45 tumor metadata": "6",
        "Paired response distribution": "Low=3; Medium=2; High=1",
    }
    for key, value in expected_fragments.items():
        fragment = f"{key}: {value}"
        add("Data boundary", key, "PASS" if fragment in summary else "FAIL", f"Expected `{fragment}` in {rel(RESULTS / 'data_audit' / 'DATA_AUDIT_SUMMARY.md')}.")
    check_text_contains(
        text,
        "Manuscript data-boundary claims",
        ["41 raw h5", "25 tumor", "16 PBMC", "74,557", "19 patients", "six strict pre/post paired"],
    )

    paired = read_csv(RESULTS / "data_audit" / "paired_patient_response_summary.csv")
    observed = {row["path_response"]: int(row["N"]) for row in paired}
    add(
        "Data boundary",
        "Paired response CSV distribution",
        "PASS" if observed == {"Low": 3, "Medium": 2, "High": 1} else "FAIL",
        f"observed={observed}; source={rel(RESULTS / 'data_audit' / 'paired_patient_response_summary.csv')}",
    )


def audit_composition():
    baseline_path = RESULTS / "pre_baseline" / "baseline_pre_composition_trend_respOrd_limma_logit.csv"
    ordinal_path = RESULTS / "dynamic_paired" / "Fig4A_composition_delta_logit_limma_respOrd_trend.csv"
    binary_path = RESULTS / "dynamic_paired" / "Fig4A_composition_delta_logit_limma_RvsNR.csv"

    baseline = read_csv(baseline_path)
    cases = [
        ("Baseline T cell logFC", row_by(baseline, "celltype", "T cell"), "logFC", -0.201, baseline_path),
        ("Baseline T cell P", row_by(baseline, "celltype", "T cell"), "P.Value", 0.663, baseline_path),
        ("Baseline T cell FDR", row_by(baseline, "celltype", "T cell"), "adj.P.Val", 0.827, baseline_path),
        ("Baseline Myeloid logFC", row_by(baseline, "celltype", "Myeloid"), "logFC", 0.135, baseline_path),
        ("Baseline Myeloid P", row_by(baseline, "celltype", "Myeloid"), "P.Value", 0.827, baseline_path),
        ("Baseline Myeloid FDR", row_by(baseline, "celltype", "Myeloid"), "adj.P.Val", 0.827, baseline_path),
    ]
    for item, row, col, expected, path in cases:
        check_numeric("Composition", item, fnum(row[col]), expected, path)

    ordinal = read_csv(ordinal_path)
    for cell, expected in {
        "Mast": (1.01, 0.00413, 0.0248),
        "Myeloid": (-0.602, 0.0703, 0.211),
        "T cell": (0.241, 0.455, 0.776),
    }.items():
        row = row_by(ordinal, "celltype", cell)
        for col, label, value in [("logFC", "logFC", expected[0]), ("P.Value", "P", expected[1]), ("adj.P.Val", "FDR", expected[2])]:
            check_numeric("Composition", f"Ordinal paired {cell} {label}", fnum(row[col]), value, ordinal_path)

    binary = read_csv(binary_path)
    for cell, expected in {
        "Myeloid": (-1.65, 0.0157, 0.0488),
        "Mast": (1.64, 0.0163, 0.0488),
        "T cell": (0.751, 0.225, 0.450),
    }.items():
        row = row_by(binary, "celltype", cell)
        for col, label, value in [("logFC", "logFC", expected[0]), ("P.Value", "P", expected[1]), ("adj.P.Val", "FDR", expected[2])]:
            check_numeric("Composition", f"Binary paired {cell} {label}", fnum(row[col]), value, binary_path)


def audit_dynamic_de_and_gsea():
    t_de_path = RESULTS / "dynamic_paired" / "Fig4B_T_cell_interaction_DE_trend.csv"
    m_de_path = RESULTS / "dynamic_paired" / "Fig4B_Myeloid_interaction_DE_trend.csv"
    t_gsea_path = RESULTS / "dynamic_paired" / "Fig4B_T_cell_GSEA_Hallmark.csv"
    m_gsea_path = RESULTS / "dynamic_paired" / "Fig4B_Myeloid_GSEA_Hallmark.csv"

    t_de = read_csv(t_de_path)
    m_de = read_csv(m_de_path)
    t_fdr05 = sum(1 for row in t_de if finite_less(row["padj"], 0.05))
    t_fdr10 = sum(1 for row in t_de if finite_less(row["padj"], 0.10))
    m_fdr05 = sum(1 for row in m_de if finite_less(row["padj"], 0.05))
    m_fdr10 = sum(1 for row in m_de if finite_less(row["padj"], 0.10))
    add("Dynamic DE", "T-cell FDR counts", "PASS" if (t_fdr05, t_fdr10) == (24, 27) else "FAIL", f"FDR<0.05={t_fdr05}; FDR<0.10={t_fdr10}; source={rel(t_de_path)}")
    add("Dynamic DE", "Myeloid FDR counts", "PASS" if (m_fdr05, m_fdr10) == (3, 4) else "FAIL", f"FDR<0.05={m_fdr05}; FDR<0.10={m_fdr10}; source={rel(m_de_path)}")

    for gene, expected_lfc, expected_fdr in [("CD9", 1.27, 0.00357), ("CD1A", 1.57, 0.0380), ("ACP5", 1.00, 0.0428)]:
        row = row_by(m_de, "gene", gene)
        check_numeric("Dynamic DE", f"Myeloid {gene} log2FC", fnum(row["log2FoldChange"]), expected_lfc, m_de_path)
        check_numeric("Dynamic DE", f"Myeloid {gene} FDR", fnum(row["padj"]), expected_fdr, m_de_path)

    t_gsea = read_csv(t_gsea_path)
    t_expected = {
        "HALLMARK_TNFA_SIGNALING_VIA_NFKB": (2.70, 1.86e-18),
        "HALLMARK_MTORC1_SIGNALING": (1.76, 2.17e-4),
        "HALLMARK_P53_PATHWAY": (1.81, 2.17e-4),
        "HALLMARK_INTERFERON_ALPHA_RESPONSE": (1.87, 7.72e-4),
        "HALLMARK_INTERFERON_GAMMA_RESPONSE": (1.66, 1.81e-3),
    }
    for pathway, expected in t_expected.items():
        row = row_by(t_gsea, "pathway", pathway)
        check_numeric("Dynamic GSEA", f"T-cell {pathway} NES", fnum(row["NES"]), expected[0], t_gsea_path)
        check_numeric("Dynamic GSEA", f"T-cell {pathway} FDR", fnum(row["padj"]), expected[1], t_gsea_path, tol=1e-2)

    m_gsea = read_csv(m_gsea_path)
    m_expected = {
        "HALLMARK_MTORC1_SIGNALING": (2.55, 1.70e-16),
        "HALLMARK_INTERFERON_GAMMA_RESPONSE": (2.24, 8.18e-11),
        "HALLMARK_TNFA_SIGNALING_VIA_NFKB": (2.24, 8.18e-11),
        "HALLMARK_INTERFERON_ALPHA_RESPONSE": (2.50, 1.97e-10),
        "HALLMARK_INFLAMMATORY_RESPONSE": (2.26, 1.98e-10),
        "HALLMARK_COMPLEMENT": (2.12, 2.12e-8),
    }
    for pathway, expected in m_expected.items():
        row = row_by(m_gsea, "pathway", pathway)
        check_numeric("Dynamic GSEA", f"Myeloid {pathway} NES", fnum(row["NES"]), expected[0], m_gsea_path)
        check_numeric("Dynamic GSEA", f"Myeloid {pathway} FDR", fnum(row["padj"]), expected[1], m_gsea_path, tol=1e-2)


def audit_leave_one_out():
    loo_path = RESULTS / "sensitivity_leave_one_out" / "LOO_model_stability_summary.csv"
    rows = read_csv(loo_path)
    for cell, expected_s, expected_overlap in [("T cell", 0.495, 0.15), ("Myeloid", 0.623, 0.26)]:
        row = next(row for row in rows if row["celltype"] == cell and row["leave_out"] == "P32")
        check_numeric("Leave-one-out", f"{cell} P32 Spearman", fnum(row["stat_spearman"]), expected_s, loo_path)
        check_numeric("Leave-one-out", f"{cell} P32 top100 overlap", fnum(row["top100_overlap"]), expected_overlap, loo_path)
        check_numeric("Leave-one-out", f"{cell} P32 top25 direction concordance", fnum(row["full_top25_direction_concordance"]), 1.0, loo_path)


def audit_abundance_exact_permutation(text: str):
    result_path = RESULTS / "sensitivity_exact_permutation" / "ABUNDANCE_EXACT_PERMUTATION_RESULTS.csv"
    null_path = RESULTS / "sensitivity_exact_permutation" / "ABUNDANCE_EXACT_PERMUTATION_NULL_DISTRIBUTIONS.csv"
    report_path = RESULTS / "sensitivity_exact_permutation" / "ABUNDANCE_EXACT_PERMUTATION_REPORT.md"
    figure_path = REBUILD_DIR / "figures" / "submission" / "ExtendedData8_submission_abundance_exact_permutation.png"
    source_paths = [
        FIGURE_SOURCE / "ExtendedData8_abundance_exact_permutation_results.csv",
        FIGURE_SOURCE / "ExtendedData8_abundance_permutation_null.csv",
    ]
    for path in [result_path, null_path, report_path, figure_path, *source_paths]:
        add(
            "Abundance exact permutation",
            path.name,
            "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL",
            f"source={rel(path)}",
        )

    rows = read_csv(result_path)

    def exact_row(analysis: str, celltype: str) -> dict[str, str]:
        return next(
            row
            for row in rows
            if row["analysis"] == analysis and row["celltype"] == celltype
        )

    ordinal_mast = exact_row("paired_delta_ordinal_unadjusted", "Mast")
    check_numeric("Abundance exact permutation", "Ordinal Mast unique assignments", fnum(ordinal_mast["n_unique_permutations"]), 60, result_path)
    check_numeric("Abundance exact permutation", "Ordinal Mast exact P", fnum(ordinal_mast["exact_p_two_sided"]), 0.0833333, result_path)
    check_numeric("Abundance exact permutation", "Ordinal Mast exact FDR", fnum(ordinal_mast["exact_bh_fdr"]), 0.5, result_path)

    binary_myeloid = exact_row("paired_delta_High_vs_Low", "Myeloid")
    check_numeric("Abundance exact permutation", "High-Low unique assignments", fnum(binary_myeloid["n_unique_permutations"]), 4, result_path)
    check_numeric("Abundance exact permutation", "High-Low Myeloid exact P", fnum(binary_myeloid["exact_p_two_sided"]), 0.25, result_path)
    check_numeric("Abundance exact permutation", "High-Low Myeloid exact FDR", fnum(binary_myeloid["exact_bh_fdr"]), 0.5, result_path)

    adjusted = [
        row for row in rows if row["analysis"] == "paired_delta_ordinal_cohort_adjusted"
    ]
    adjusted_ok = (
        len(adjusted) == 6
        and all(int(row["n_unique_permutations"]) == 18 for row in adjusted)
        and all(fnum(row["exact_bh_fdr"]) >= 0.05 for row in adjusted)
    )
    add(
        "Abundance exact permutation",
        "Cohort-adjusted exact inference",
        "PASS" if adjusted_ok else "FAIL",
        f"celltypes={len(adjusted)}; all_18_assignments={all(int(row['n_unique_permutations']) == 18 for row in adjusted)}; none_exact_FDR_significant={all(fnum(row['exact_bh_fdr']) >= 0.05 for row in adjusted)}; source={rel(result_path)}",
    )
    check_text_contains(
        text,
        "Abundance exact-permutation manuscript boundary",
        [
            "no broad compartment survived exhaustive patient-label permutation",
            "exact P = 0.0833",
            "exact P = 0.250",
            "descriptive candidates, not statistically confirmed supporting signals",
        ],
    )


def audit_discovery_cohort_sensitivity(text: str):
    sensitivity_dir = RESULTS / "sensitivity_cohort_adjusted_pseudobulk"
    key_path = sensitivity_dir / "KEY_PATHWAY_MODEL_COMPARISON.csv"
    permutation_path = sensitivity_dir / "KEY_PATHWAY_STRATIFIED_EXACT_PERMUTATION.csv"
    report_path = sensitivity_dir / "DISCOVERY_COHORT_ADJUSTED_PSEUDOBULK_REPORT.md"
    figure_path = REBUILD_DIR / "figures" / "submission" / "ExtendedData9_submission_discovery_cohort_sensitivity.png"
    source_paths = [
        FIGURE_SOURCE / "ExtendedData9_key_pathway_model_comparison.csv",
        FIGURE_SOURCE / "ExtendedData9_key_pathway_exact_permutation.csv",
    ]
    for path in [key_path, permutation_path, report_path, figure_path, *source_paths]:
        add(
            "Discovery cohort sensitivity",
            path.name,
            "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL",
            f"source={rel(path)}",
        )

    key_rows = read_csv(key_path)
    key_ok = (
        len(key_rows) == 12
        and all(fnum(row["adjusted_NES"]) > 0 for row in key_rows)
        and all(fnum(row["adjusted_FDR"]) < 0.05 for row in key_rows)
    )
    add(
        "Discovery cohort sensitivity",
        "Adjusted key-pathway direction and FDR",
        "PASS" if key_ok else "FAIL",
        f"pathways={len(key_rows)}; all_adjusted_NES_positive={all(fnum(row['adjusted_NES']) > 0 for row in key_rows)}; all_adjusted_FDR_lt_0.05={all(fnum(row['adjusted_FDR']) < 0.05 for row in key_rows)}; source={rel(key_path)}",
    )

    permutation_rows = read_csv(permutation_path)
    permutation_ok = (
        len(permutation_rows) == 12
        and all(int(row["n_paired_patients"]) == 6 for row in permutation_rows)
        and all(int(row["n_unique_stratified_assignments"]) == 18 for row in permutation_rows)
        and all(fnum(row["exact_p_two_sided"]) >= 0.05 for row in permutation_rows)
        and all(fnum(row["exact_bh_fdr"]) >= 0.05 for row in permutation_rows)
    )
    add(
        "Discovery cohort sensitivity",
        "Strict-pair exact pathway inference",
        "PASS" if permutation_ok else "FAIL",
        f"pathways={len(permutation_rows)}; all_n_pairs_6={all(int(row['n_paired_patients']) == 6 for row in permutation_rows)}; all_18_assignments={all(int(row['n_unique_stratified_assignments']) == 18 for row in permutation_rows)}; none_exact_P_significant={all(fnum(row['exact_p_two_sided']) >= 0.05 for row in permutation_rows)}; source={rel(permutation_path)}",
    )
    check_text_contains(
        text,
        "Discovery cohort-sensitivity manuscript boundary",
        [
            "Spearman rho = 0.875 for T cells and 0.895 for myeloid cells",
            "None of the 12 lineage-pathway tests was significant",
            "18 possible assignments",
            "did not provide confirmatory pathway significance",
            "Extended Data Fig. 9",
        ],
    )


def audit_external_validation(text: str):
    stats_path = RESULTS / "external_validation" / "GSE123813_paired_delta_stats.csv"
    cluster_path = RESULTS / "external_validation" / "GSE123813_cluster_composition_delta_stats.csv"
    interp_path = RESULTS / "external_validation" / "GSE123813_validation_interpretation.md"

    for path in [stats_path, cluster_path, interp_path]:
        add("External validation", path.name, "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL", f"source={rel(path)}")

    stats_rows = read_csv(stats_path)
    t_anchor = row_by(stats_rows, "signature", "T_DE_FDR05_positive")
    check_numeric("External validation", "T DE module n paired", fnum(t_anchor["n_paired_patients"]), 11, stats_path)
    check_numeric("External validation", "T DE module mean delta", fnum(t_anchor["mean_post_minus_pre"]), -0.0070415, stats_path)
    check_numeric("External validation", "T DE module P", fnum(t_anchor["paired_t_pvalue"]), 0.90637, stats_path)
    check_numeric("External validation", "T DE module FDR", fnum(t_anchor["paired_t_fdr"]), 0.95353, stats_path)

    m_ifna = row_by(stats_rows, "signature", "M_LE_INTERFERON_ALPHA_RESPONSE")
    check_numeric("External validation", "Myeloid IFNA module n paired", fnum(m_ifna["n_paired_patients"]), 6, stats_path)
    check_numeric("External validation", "Myeloid IFNA module mean delta", fnum(m_ifna["mean_post_minus_pre"]), 0.070628, stats_path)
    check_numeric("External validation", "Myeloid IFNA module P", fnum(m_ifna["paired_t_pvalue"]), 0.12756, stats_path)
    check_numeric("External validation", "Myeloid IFNA module FDR", fnum(m_ifna["paired_t_fdr"]), 0.81096, stats_path)

    cluster_rows = read_csv(cluster_path)
    cd8_ex = row_by(cluster_rows, "cluster", "CD8_ex_T_cells")
    check_numeric("External validation", "CD8 exhausted cluster mean logit delta", fnum(cd8_ex["mean_logit_post_minus_pre"]), 0.87118, cluster_path)
    check_numeric("External validation", "CD8 exhausted positive patients", fnum(cd8_ex["positive_delta_patients"]), 9, cluster_path)
    check_numeric("External validation", "CD8 exhausted Wilcoxon FDR", fnum(cd8_ex["wilcoxon_fdr"]), 0.30469, cluster_path)

    check_text_contains(
        text,
        "External validation manuscript boundary",
        ["GSE123813", "boundary test", "rather than as an external response-prediction validation cohort", "fixed cross-disease"],
    )


def audit_gse281729_robustness(text: str):
    validation_dir = VALIDATION / "GSE281729_bulk_module_validation"
    model_path = validation_dir / "GSE281729_ROBUST_RESPONSE_MODELS.csv"
    loo_path = validation_dir / "GSE281729_LEAVE_ONE_PATIENT_SUMMARY.csv"
    influence_path = validation_dir / "GSE281729_COOKS_DISTANCE_SUMMARY.csv"
    coverage_path = validation_dir / "GSE281729_LOCKED_MODULE_GENE_COVERAGE.csv"
    source_paths = [
        FIGURE_SOURCE / "ExtendedData7_model_comparison_source.csv",
        FIGURE_SOURCE / "ExtendedData7_leave_one_out_source.csv",
        FIGURE_SOURCE / "ExtendedData7_influence_source.csv",
        FIGURE_SOURCE / "ExtendedData7_gene_coverage_source.csv",
    ]
    for path in [model_path, loo_path, influence_path, coverage_path, *source_paths]:
        add(
            "GSE281729 robustness",
            path.name,
            "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL",
            f"source={rel(path)}",
        )

    leading = {
        "M_LE_INTERFERON_ALPHA_RESPONSE",
        "T_LE_INTERFERON_ALPHA_RESPONSE",
        "M_LE_INTERFERON_GAMMA_RESPONSE",
        "M_LE_union_core",
        "T_LE_INTERFERON_GAMMA_RESPONSE",
        "M_LE_MTORC1_SIGNALING",
        "T_LE_MTORC1_SIGNALING",
        "T_LE_union_core",
    }
    robust = [
        row
        for row in read_csv(model_path)
        if row["model"] == "adjusted_hpv_second_drug_HC3" and row["signature"] in leading
    ]
    robust_ok = (
        len(robust) == 8
        and all(fnum(row["coef"]) < 0 for row in robust)
        and all(close(fnum(row["fdr"]), 0.0340, rel_tol=0, abs_tol=1e-4) for row in robust)
    )
    add(
        "GSE281729 robustness",
        "Adjusted HC3 leading modules",
        "PASS" if robust_ok else "FAIL",
        f"modules={len(robust)}; all_negative={all(fnum(row['coef']) < 0 for row in robust)}; all_FDR_0.0340={all(close(fnum(row['fdr']), 0.0340, rel_tol=0, abs_tol=1e-4) for row in robust)}; source={rel(model_path)}",
    )

    loo = [
        row
        for row in read_csv(loo_path)
        if row["model"] == "adjusted_hpv_second_drug" and row["signature"] in leading
    ]
    loo_ok = len(loo) == 8 and sum(int(row["n_refits"]) for row in loo) == 240 and all(fnum(row["negative_refit_fraction"]) == 1 for row in loo)
    add(
        "GSE281729 robustness",
        "Adjusted leave-one-patient stability",
        "PASS" if loo_ok else "FAIL",
        f"modules={len(loo)}; total_refits={sum(int(row['n_refits']) for row in loo)}; all_negative={all(fnum(row['negative_refit_fraction']) == 1 for row in loo)}; source={rel(loo_path)}",
    )

    influence = [
        row
        for row in read_csv(influence_path)
        if row["model"] == "adjusted_hpv_second_drug" and row["signature"] in leading
    ]
    influence_ok = len(influence) == 8 and all(int(row["n_above_threshold"]) > 0 for row in influence) and all(close(fnum(row["threshold_4_over_n"]), 4 / 30, rel_tol=0, abs_tol=1e-9) for row in influence)
    add(
        "GSE281729 robustness",
        "Cook distance boundary",
        "PASS" if influence_ok else "FAIL",
        f"modules={len(influence)}; all_have_flag={all(int(row['n_above_threshold']) > 0 for row in influence)}; threshold=4/30; source={rel(influence_path)}",
    )

    coverage = [row for row in read_csv(coverage_path) if row["signature"] in leading]
    fractions = [fnum(row["n_genes_present_in_GSE281729"]) / fnum(row["n_genes_defined"]) for row in coverage]
    coverage_ok = len(coverage) == 8 and close(min(fractions), 84 / 87, rel_tol=0, abs_tol=1e-9) and close(max(fractions), 1, rel_tol=0, abs_tol=1e-9)
    add(
        "GSE281729 robustness",
        "Locked-module gene coverage",
        "PASS" if coverage_ok else "FAIL",
        f"modules={len(coverage)}; range={100 * min(fractions):.1f}-{100 * max(fractions):.1f}%; source={rel(coverage_path)}",
    )

    check_text_contains(
        text,
        "GSE281729 robustness manuscript claims",
        [
            "HC3 heteroskedasticity-consistent covariance",
            "All 30 leave-one-patient refits per module retained negative adjusted slopes",
            "Cook's-distance threshold of 4/n",
            "Module gene coverage ranged from 96.6% to 100%",
        ],
    )


def audit_gse301741_boundary(text: str):
    validation_dir = VALIDATION / "GSE301741_lineage_aware_validation"
    response_path = validation_dir / "GSE301741_LINEAGE_AWARE_RESPONSE_TESTS.csv"
    pair_qc_path = validation_dir / "GSE301741_LINEAGE_AWARE_PAIR_CELL_QC.csv"
    figure_stem = REBUILD_DIR / "figures" / "submission" / "ExtendedData10_submission_gse301741_boundary"
    source_stem = FIGURE_SOURCE / "ExtendedData10_submission_gse301741_boundary"
    asset_paths = [
        response_path,
        pair_qc_path,
        figure_stem.with_suffix(".png"),
        figure_stem.with_suffix(".pdf"),
        figure_stem.with_suffix(".svg"),
        source_stem.with_name(source_stem.name + "_source_data.xlsx"),
        source_stem.with_name(source_stem.name + "_panel_a_lineage_composition.csv"),
        source_stem.with_name(source_stem.name + "_panel_b_pair_cell_gate.csv"),
        source_stem.with_name(source_stem.name + "_panel_c_patient_module_delta.csv"),
        source_stem.with_name(source_stem.name + "_panel_d_exact_permutation.csv"),
    ]
    for path in asset_paths:
        add(
            "GSE301741 boundary",
            path.name,
            "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL",
            f"source={rel(path)}",
        )

    response_rows = read_csv(response_path)
    t_rows = [row for row in response_rows if row["target_lineage"] == "T_cell"]
    myeloid_rows = [row for row in response_rows if row["target_lineage"] == "Myeloid"]
    t_ok = (
        len(t_rows) == 7
        and all(int(row["n_responder"]) == 3 and int(row["n_non_responder"]) == 4 for row in t_rows)
        and all(fnum(row["exact_permutation_p"]) >= 0.6857 for row in t_rows)
        and all(fnum(row["permutation_fdr"]) == 1 for row in t_rows)
    )
    add(
        "GSE301741 boundary",
        "Exact T-cell response inference",
        "PASS" if t_ok else "FAIL",
        f"modules={len(t_rows)}; all_3v4={all(int(row['n_responder']) == 3 and int(row['n_non_responder']) == 4 for row in t_rows)}; min_exact_P={min(fnum(row['exact_permutation_p']) for row in t_rows):.6f}; all_FDR_1={all(fnum(row['permutation_fdr']) == 1 for row in t_rows)}; source={rel(response_path)}",
    )
    myeloid_ok = (
        len(myeloid_rows) > 0
        and all(int(row["n_responder"]) == 0 for row in myeloid_rows)
        and all(row["inference_status"] == "descriptive_only_insufficient_response_groups" for row in myeloid_rows)
    )
    add(
        "GSE301741 boundary",
        "Myeloid response comparison remains non-estimable",
        "PASS" if myeloid_ok else "FAIL",
        f"modules={len(myeloid_rows)}; all_responder_n_0={all(int(row['n_responder']) == 0 for row in myeloid_rows)}; source={rel(response_path)}",
    )
    check_text_contains(
        text,
        "GSE301741 manuscript and figure boundary",
        [
            "Extended Data Fig. 10",
            "all exact P >= 0.686",
            "all FDR = 1",
            "precluding a myeloid response comparison",
            "did not independently replicate the locked T-cell response association",
        ],
    )


def audit_external_tcr_validation(text: str):
    delta_path = RESULTS / "external_tcr_validation" / "GSE123813_TCR_paired_delta_stats.csv"
    turnover_path = RESULTS / "external_tcr_validation" / "GSE123813_TCR_pair_turnover.csv"
    summary_path = RESULTS / "external_tcr_validation" / "GSE123813_TCR_turnover_summary.csv"
    interp_path = RESULTS / "external_tcr_validation" / "GSE123813_TCR_validation_interpretation.md"

    for path in [delta_path, turnover_path, summary_path, interp_path]:
        add("External TCR validation", path.name, "PASS" if path.exists() and path.stat().st_size > 0 else "FAIL", f"source={rel(path)}")

    delta_rows = read_csv(delta_path)
    all_cd8 = next(row for row in delta_rows if row["disease_group"] == "ALL" and row["metric"] == "cd8_ex_fraction")
    check_numeric("External TCR validation", "ALL CD8 exhausted n paired", fnum(all_cd8["n_paired_patients"]), 15, delta_path)
    check_numeric("External TCR validation", "ALL CD8 exhausted mean delta", fnum(all_cd8["mean_post_minus_pre"]), 0.042313, delta_path)
    check_numeric("External TCR validation", "ALL CD8 exhausted Wilcoxon P", fnum(all_cd8["wilcoxon_pvalue"]), 0.035339, delta_path)
    check_numeric("External TCR validation", "ALL CD8 exhausted Wilcoxon FDR", fnum(all_cd8["wilcoxon_fdr"]), 0.424072, delta_path)

    bcc_cd8 = next(row for row in delta_rows if row["disease_group"] == "BCC" and row["metric"] == "cd8_ex_fraction")
    check_numeric("External TCR validation", "BCC CD8 exhausted mean delta", fnum(bcc_cd8["mean_post_minus_pre"]), 0.043488, delta_path)
    check_numeric("External TCR validation", "BCC CD8 exhausted Wilcoxon P", fnum(bcc_cd8["wilcoxon_pvalue"]), 0.013672, delta_path)
    check_numeric("External TCR validation", "BCC CD8 exhausted Wilcoxon FDR", fnum(bcc_cd8["wilcoxon_fdr"]), 0.328125, delta_path)

    summary_rows = read_csv(summary_path)
    all_post_new = next(row for row in summary_rows if row["disease_group"] == "ALL" and row["metric"] == "post_new_cell_fraction")
    all_mh = next(row for row in summary_rows if row["disease_group"] == "ALL" and row["metric"] == "morisita_horn_overlap")
    bcc_post_new = next(row for row in summary_rows if row["disease_group"] == "BCC" and row["metric"] == "post_new_cell_fraction")
    check_numeric("External TCR validation", "ALL post-new clonotype median", fnum(all_post_new["median"]), 0.758333, summary_path)
    check_numeric("External TCR validation", "BCC post-new clonotype median", fnum(bcc_post_new["median"]), 0.792952, summary_path)
    check_numeric("External TCR validation", "ALL Morisita-Horn median", fnum(all_mh["median"]), 0.405824, summary_path)

    check_text_contains(
        text,
        "External TCR validation manuscript boundary",
        ["orthogonal TCR", "post-new clonotype", "Morisita-Horn", "not response-depth validation"],
    )


def audit_workbooks_and_bundle():
    expected_workbooks = [
        "Figure1_source_data.xlsx",
        "Figure2_source_data.xlsx",
        "ExtendedData1_source_data.xlsx",
        "ExtendedData2_source_data.xlsx",
        "ExtendedData3_source_data.xlsx",
        "ExtendedData4_source_data.xlsx",
        "ExtendedData5_source_data.xlsx",
        "ExtendedData6_source_data.xlsx",
        "ExtendedData10_submission_gse301741_boundary_source_data.xlsx",
    ]
    for name in expected_workbooks:
        path = FIGURE_SOURCE / name
        try:
            wb = load_workbook(path, read_only=False, data_only=False)
            sheets = wb.sheetnames
            wb.close()
            add("Workbook integrity", name, "PASS", f"{len(sheets)} sheets; sha256={sha256(path)[:16]}; source={rel(path)}")
        except Exception as exc:  # noqa: BLE001
            add("Workbook integrity", name, "FAIL", str(exc))

    supp = TABLE_DIR / "Supplementary_Tables.xlsx"
    try:
        wb = load_workbook(supp, read_only=False, data_only=False)
        count = len(wb.sheetnames)
        wb.close()
        add("Workbook integrity", "Supplementary_Tables.xlsx", "PASS" if count == 61 else "FAIL", f"{count} sheets; source={rel(supp)}")
    except Exception as exc:  # noqa: BLE001
        add("Workbook integrity", "Supplementary_Tables.xlsx", "FAIL", str(exc))

    if ZIP_PATH.exists():
        with zipfile.ZipFile(ZIP_PATH, "r") as zf:
            bad = zf.testzip()
            entries = [name for name in zf.namelist() if not name.endswith("/")]
        add("Bundle integrity", "Submission zip", "PASS" if bad is None and len(entries) >= 130 else "FAIL", f"entries={len(entries)}; bad={bad}; source={rel(ZIP_PATH)}")
    else:
        add(
            "Bundle integrity",
            "Submission zip",
            "WARN",
            f"Missing {rel(ZIP_PATH)}; expected after content-first cleanup before final packaging.",
        )


def audit_references_and_target_texts():
    ref_path = MANUSCRIPT_DIR / "REFERENCE_MASTER.csv"
    if not ref_path.exists():
        ref_path = MANUSCRIPT_DIR / "REFERENCE_MASTER.csv"
    refs = read_csv(ref_path)
    doi_missing = [row["short_key"] for row in refs if not row.get("doi")]
    expected_refs = 20 if ref_path.name.endswith(".csv") else 17
    add(
        "References",
        "Reference master DOI/source URLs",
        "PASS" if len(refs) == expected_refs and not doi_missing else "FAIL",
        f"references={len(refs)}; expected={expected_refs}; missing_doi={doi_missing}; source={rel(ref_path)}",
    )

    word_count_path = MANUSCRIPT_DIR / "TARGET_SPECIFIC_TEXT_WORDCOUNTS.csv"
    if word_count_path.exists():
        word_counts = read_csv(word_count_path)
        bad = [row["text_block"] for row in word_counts if row["status"] != "PASS"]
        add("Target-specific text", "Word-count checks", "PASS" if not bad else "FAIL", f"non-pass={bad}; source={rel(word_count_path)}")
    else:
        add(
            "Target-specific text",
            "Word-count checks",
            "WARN",
            f"Skipped because {rel(word_count_path)} was removed during content-first cleanup; rerun target adaptation after journal selection.",
        )


def write_outputs():
    with OUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["area", "item", "status", "detail"])
        writer.writeheader()
        writer.writerows([check.__dict__ for check in checks])

    counts = {status: sum(1 for check in checks if check.status == status) for status in ["PASS", "WARN", "FAIL"]}
    lines = [
        "# Result Consistency Audit",
        "",
        f"Date: {date.today().isoformat()}",
        "",
        "## Summary",
        "",
        f"- PASS: {counts['PASS']}",
        f"- WARN: {counts['WARN']}",
        f"- FAIL: {counts['FAIL']}",
        f"- Manuscript audited: `{rel(MANUSCRIPT)}`",
        "",
        "This audit checks whether key manuscript claims remain traceable to recomputed result tables, source-data workbooks and the current submission bundle.",
        "",
        "## Checks",
        "",
        "| Area | Item | Status | Detail |",
        "|---|---|---:|---|",
    ]
    for check in checks:
        lines.append(f"| {check.area} | `{check.item}` | {check.status} | {check.detail.replace('|', '\\|')} |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    text = MANUSCRIPT.read_text(encoding="utf-8")
    audit_data_boundary(text)
    audit_composition()
    audit_dynamic_de_and_gsea()
    audit_leave_one_out()
    audit_abundance_exact_permutation(text)
    audit_discovery_cohort_sensitivity(text)
    audit_external_validation(text)
    audit_gse281729_robustness(text)
    audit_gse301741_boundary(text)
    audit_external_tcr_validation(text)
    audit_workbooks_and_bundle()
    audit_references_and_target_texts()
    write_outputs()
    print(OUT_MD)
    print(OUT_CSV)


if __name__ == "__main__":
    main()

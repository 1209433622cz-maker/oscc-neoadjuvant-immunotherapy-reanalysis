#!/usr/bin/env python
"""Rebuild valid figure source-data XLSX workbooks from per-panel CSV files."""

from __future__ import annotations

import argparse
import csv
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_PATH = Path(__file__).resolve()
REBUILD_DIR = SCRIPT_PATH.parents[1]
SOURCE_DATA_DIR = REBUILD_DIR / "figures" / "submission" / "source_data"

DEFAULT_STEMS = [
    "Figure1",
    "Figure2",
    "ExtendedData1",
    "ExtendedData2",
    "ExtendedData3",
    "ExtendedData4",
    "ExtendedData5",
    "ExtendedData6",
]


def coerce_value(value: str):
    if value == "":
        return None
    if value.upper() in {"NA", "NAN", "NULL"}:
        return None
    if re.fullmatch(r"[-+]?\d+", value):
        try:
            return int(value)
        except ValueError:
            return value
    if re.fullmatch(r"[-+]?(\d+\.\d*|\d*\.\d+)([eE][-+]?\d+)?", value) or re.fullmatch(
        r"[-+]?\d+[eE][-+]?\d+", value
    ):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def safe_sheet_name(raw: str, used: set[str]) -> str:
    name = re.sub(r"[\[\]\*\?/\\:]", "_", raw)
    name = re.sub(r"\s+", "_", name).strip("_")
    name = name[:31] or "Sheet"
    candidate = name
    idx = 2
    while candidate in used:
        suffix = f"_{idx}"
        candidate = name[: 31 - len(suffix)] + suffix
        idx += 1
    used.add(candidate)
    return candidate


def table_name(stem: str, sheet_name: str, index: int) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", f"{stem}_{sheet_name}_{index}")
    if not re.match(r"[A-Za-z_]", base):
        base = f"T_{base}"
    return base[:200]


def csv_files_for_stem(stem: str) -> list[Path]:
    return sorted(
        path
        for path in SOURCE_DATA_DIR.glob(f"{stem}_*.csv")
        if not path.name.endswith("_source_data.csv")
    )


def append_csv_sheet(wb: Workbook, csv_path: Path, sheet_name: str, stem: str, index: int):
    ws = wb.create_sheet(sheet_name)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            ws.append([coerce_value(value) for value in row])

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 2 and ws.max_column >= 1:
        tab = Table(displayName=table_name(stem, sheet_name, index), ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 55)


def build_workbook(stem: str) -> Path:
    csv_paths = csv_files_for_stem(stem)
    if not csv_paths:
        raise FileNotFoundError(f"No CSV files found for stem {stem} in {SOURCE_DATA_DIR}")

    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    rows = [
        ("source_workbook", f"{stem}_source_data.xlsx"),
        ("generated_on", date.today().isoformat()),
        ("generator", "03_rebuild/analysis/11_rebuild_figure_source_workbooks.py"),
        ("source_csv_count", len(csv_paths)),
        ("note", "Workbook rebuilt from per-panel CSV files to avoid broken OOXML drawing references."),
    ]
    for row in rows:
        readme.append(row)
    for cell in readme["A"]:
        cell.font = Font(bold=True)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 90

    used = {"README"}
    for idx, csv_path in enumerate(csv_paths, start=1):
        label = csv_path.stem.removeprefix(f"{stem}_")
        sheet_name = safe_sheet_name(label, used)
        append_csv_sheet(wb, csv_path, sheet_name, stem, idx)

    out = SOURCE_DATA_DIR / f"{stem}_source_data.xlsx"
    wb.save(out)

    # Reopen immediately so broken packages fail during generation, not submission.
    check = load_workbook(out, read_only=False, data_only=False)
    check.close()
    return out


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stem", action="append", help="Workbook stem to rebuild; may be repeated.")
    return parser.parse_args()


def main():
    args = parse_args()
    stems = args.stem or DEFAULT_STEMS
    for stem in stems:
        out = build_workbook(stem)
        print(out)


if __name__ == "__main__":
    main()

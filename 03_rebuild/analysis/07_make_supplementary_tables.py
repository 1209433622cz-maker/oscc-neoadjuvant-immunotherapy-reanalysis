#!/usr/bin/env python
"""Assemble submission-ready supplementary tables from rebuilt result CSV files."""

from __future__ import annotations

import csv
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


SCRIPT_PATH = Path(__file__).resolve()
REBUILD_DIR = SCRIPT_PATH.parents[1]
WORKSPACE_DIR = REBUILD_DIR.parent
RESULTS_DIR = REBUILD_DIR / "results"
TABLES_DIR = REBUILD_DIR / "tables" / "submission"
CSV_DIR = TABLES_DIR / "csv"
OUT_XLSX = TABLES_DIR / "Supplementary_Tables.xlsx"
TABLES_DIR.mkdir(parents=True, exist_ok=True)
CSV_DIR.mkdir(parents=True, exist_ok=True)


@dataclass(frozen=True)
class TableSpec:
    sheet: str
    rel_path: str
    description: str
    matrix_gene_index: bool = False


TABLE_SPECS = [
    TableSpec("S1_raw_h5_manifest", "data_audit/raw_h5_manifest.csv", "Raw h5 manifest retained after workspace cleanup."),
    TableSpec("S1_patient_presence", "data_audit/cd45_tumor_patient_timepoint_presence.csv", "CD45 tumor patient pre/post availability and response labels."),
    TableSpec("S1_patient_response", "data_audit/cd45_tumor_patient_response_table.csv", "Patient-level response and cohort labels."),
    TableSpec("S1_celltype_counts", "data_audit/cd45_tumor_celltype_counts.csv", "Overall CD45 tumor immune-cell type counts."),
    TableSpec("S1_patient_cell_counts", "data_audit/cd45_tumor_patient_celltype_counts.csv", "Patient/timepoint/cell-type counts."),
    TableSpec("S2_baseline_comp", "pre_baseline/baseline_pre_composition_trend_respOrd_limma_logit.csv", "Baseline pre-treatment composition ordinal response model."),
    TableSpec("S2_dynamic_comp_ord", "dynamic_paired/Fig4A_composition_delta_logit_limma_respOrd_trend.csv", "Paired post-pre composition ordinal response model."),
    TableSpec("S2_dynamic_comp_RvsNR", "dynamic_paired/Fig4A_composition_delta_logit_limma_RvsNR.csv", "Paired post-pre composition binary R/NR sensitivity model."),
    TableSpec("S3_baseline_T_DE", "pre_baseline/Baseline_T_cell_DE_trend_respOrd.csv", "Baseline T-cell pseudobulk response-trend DE."),
    TableSpec("S3_baseline_T_GSEA", "pre_baseline/Baseline_T_cell_GSEA_Hallmark.csv", "Baseline T-cell Hallmark GSEA."),
    TableSpec("S3_baseline_M_DE", "pre_baseline/Baseline_Myeloid_DE_trend_respOrd.csv", "Baseline myeloid pseudobulk response-trend DE."),
    TableSpec("S3_baseline_M_GSEA", "pre_baseline/Baseline_Myeloid_GSEA_Hallmark.csv", "Baseline myeloid Hallmark GSEA."),
    TableSpec("S4_dynamic_T_DE", "dynamic_paired/Fig4B_T_cell_interaction_DE_trend.csv", "Dynamic T-cell post-by-response interaction DE."),
    TableSpec("S4_dynamic_T_GSEA", "dynamic_paired/Fig4B_T_cell_GSEA_Hallmark.csv", "Dynamic T-cell Hallmark GSEA."),
    TableSpec("S4_T_signature_delta", "dynamic_paired/Fig4B_T_cell_SignatureDelta_source.csv", "T-cell signature deltas in paired patients."),
    TableSpec("S4_T_delta_matrix", "dynamic_paired/Fig4B_Tcell_delta_matrix.csv", "T-cell top-gene post-pre delta matrix.", True),
    TableSpec("S5_dynamic_M_DE", "dynamic_paired/Fig4B_Myeloid_interaction_DE_trend.csv", "Dynamic myeloid post-by-response interaction DE."),
    TableSpec("S5_dynamic_M_GSEA", "dynamic_paired/Fig4B_Myeloid_GSEA_Hallmark.csv", "Dynamic myeloid Hallmark GSEA."),
    TableSpec("S5_M_signature_delta", "dynamic_paired/Fig4B_Myeloid_SignatureDelta_source.csv", "Myeloid signature deltas in paired patients."),
    TableSpec("S5_M_delta_matrix", "dynamic_paired/Fig4B_Myeloid_delta_matrix.csv", "Myeloid top-gene post-pre delta matrix.", True),
    TableSpec("S6_LOO_stability", "sensitivity_leave_one_out/LOO_model_stability_summary.csv", "Leave-one-patient model stability summary."),
    TableSpec("S6_LOO_key_pathways", "sensitivity_leave_one_out/LOO_key_pathway_NES.csv", "Leave-one-patient key pathway NES."),
    TableSpec("S6_LOO_direction", "sensitivity_leave_one_out/LOO_top25_gene_direction_concordance.csv", "Leave-one-patient top-25 direction concordance."),
    TableSpec("S7_external_manifest", "external_validation/GSE123813_gene_set_manifest.csv", "External GSE123813 validation gene-set manifest."),
    TableSpec("S7_external_scores", "external_validation/GSE123813_patient_signature_scores.csv", "External GSE123813 patient-level signature scores."),
    TableSpec("S7_external_delta_stats", "external_validation/GSE123813_paired_delta_stats.csv", "External GSE123813 paired post-pre signature statistics."),
    TableSpec("S7_external_cluster_comp", "external_validation/GSE123813_patient_cluster_composition.csv", "External GSE123813 patient-level cluster composition."),
    TableSpec("S7_external_cluster_delta", "external_validation/GSE123813_cluster_composition_delta_stats.csv", "External GSE123813 paired cluster-composition statistics."),
    TableSpec("S8_TCR_metadata", "external_tcr_validation/GSE123813_TCR_metadata_summary.csv", "External GSE123813 BCC/SCC TCR validation metadata summary."),
    TableSpec("S8_TCR_sample_metrics", "external_tcr_validation/GSE123813_TCR_sample_metrics.csv", "External GSE123813 TCR sample-level clonotype and state metrics."),
    TableSpec("S8_TCR_delta_stats", "external_tcr_validation/GSE123813_TCR_paired_delta_stats.csv", "External GSE123813 paired TCR metric deltas."),
    TableSpec("S8_TCR_turnover", "external_tcr_validation/GSE123813_TCR_pair_turnover.csv", "External GSE123813 pairwise pre/post TCR turnover."),
    TableSpec("S8_TCR_turnover_summary", "external_tcr_validation/GSE123813_TCR_turnover_summary.csv", "External GSE123813 TCR turnover summary statistics."),
]


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\*\?/\\:]", "_", name)
    return cleaned[:31]


def load_table(spec: TableSpec) -> pd.DataFrame:
    source = RESULTS_DIR / spec.rel_path
    if not source.exists():
        raise FileNotFoundError(f"Required file not found for {spec.sheet}: {source}")
    df = pd.read_csv(source)
    if spec.matrix_gene_index:
        first = df.columns[0]
        df = df.rename(columns={first: "gene"})
    return df


def write_csv(df: pd.DataFrame, sheet: str) -> Path:
    out = CSV_DIR / f"{re.sub(r'[^A-Za-z0-9_]+', '_', sheet)}.csv"
    df.to_csv(out, index=False, quoting=csv.QUOTE_MINIMAL)
    return out


def value_for_excel(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def add_dataframe_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(safe_sheet_name(name))
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([value_for_excel(v) for v in row])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    sample = df.head(200).astype(str)
    for idx, col in enumerate(df.columns, start=1):
        values = [str(col)] + sample[col].fillna("").tolist()
        width = min(max(len(v) for v in values) + 2, 48)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    print(f"Workspace: {WORKSPACE_DIR}")
    print(f"Results:   {RESULTS_DIR}")
    print(f"Tables:    {TABLES_DIR}")

    tables: dict[str, pd.DataFrame] = {}
    manifest_rows = []

    for spec in TABLE_SPECS:
        df = load_table(spec)
        tables[spec.sheet] = df
        csv_path = write_csv(df, spec.sheet)
        manifest_rows.append(
            {
                "sheet": spec.sheet,
                "source_file": str(RESULTS_DIR / spec.rel_path),
                "csv_file": str(csv_path),
                "n_rows": len(df),
                "n_cols": len(df.columns),
                "description": spec.description,
            }
        )

    manifest = pd.DataFrame(manifest_rows).sort_values("sheet").reset_index(drop=True)
    readme_csv = CSV_DIR / "README.csv"
    manifest.to_csv(readme_csv, index=False)

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_dataframe_sheet(wb, "README", manifest)
    for sheet, df in tables.items():
        add_dataframe_sheet(wb, sheet, df)
    wb.save(OUT_XLSX)

    summary_path = TABLES_DIR / "SUPPLEMENTARY_TABLES_MANIFEST.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["generated_at_note", "workbook", "csv_directory", "n_tables"])
        writer.writeheader()
        writer.writerow(
            {
                "generated_at_note": "generated by 07_make_supplementary_tables.py",
                "workbook": str(OUT_XLSX),
                "csv_directory": str(CSV_DIR),
                "n_tables": len(tables),
            }
        )

    print(f"Done. Supplementary tables written to: {OUT_XLSX}")


if __name__ == "__main__":
    main()

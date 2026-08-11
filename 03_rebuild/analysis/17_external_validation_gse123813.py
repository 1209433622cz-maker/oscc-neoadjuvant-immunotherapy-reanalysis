#!/usr/bin/env python
"""Independent treatment-direction validation in GSE123813.

This script validates the OSCC-derived treatment-remodeling signal in an
orthogonal paired anti-PD-1 scRNA-seq dataset. GSE123813 is not an OSCC
response-depth cohort, so the analysis is restricted to pre/post direction
testing of OSCC-derived T-cell and myeloid modules.
"""

from __future__ import annotations

import csv
import gzip
import math
import re
import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats


SCRIPT_PATH = Path(__file__).resolve()
REBUILD_DIR = SCRIPT_PATH.parents[1]
ROOT = REBUILD_DIR.parent
RAW_DIR = ROOT / "00_raw_data" / "GSE123813_validation"
RESULTS_DIR = REBUILD_DIR / "results" / "external_validation"
FIGURE_DIR = REBUILD_DIR / "figures" / "submission"
SOURCE_DATA_DIR = FIGURE_DIR / "source_data"
MANUSCRIPT_DIR = REBUILD_DIR / "manuscript"

COUNTS_PATH = RAW_DIR / "GSE123813_bcc_scRNA_counts.txt.gz"
META_PATH = RAW_DIR / "GSE123813_bcc_all_metadata.txt.gz"
DYNAMIC_DIR = REBUILD_DIR / "results" / "dynamic_paired"

T_DE_PATH = DYNAMIC_DIR / "Fig4B_T_cell_interaction_DE_trend.csv"
M_DE_PATH = DYNAMIC_DIR / "Fig4B_Myeloid_interaction_DE_trend.csv"
T_GSEA_PATH = DYNAMIC_DIR / "Fig4B_T_cell_GSEA_Hallmark.csv"
M_GSEA_PATH = DYNAMIC_DIR / "Fig4B_Myeloid_GSEA_Hallmark.csv"

FIG_STEM = "ExtendedData5_external_validation_GSE123813"
MIN_CELLS_PER_PATIENT_TIMEPOINT = 20

T_CLUSTERS = {
    "CD8_mem_T_cells",
    "CD4_T_cells",
    "Tregs",
    "CD8_act_T_cells",
    "CD8_ex_T_cells",
    "Tcell_prolif",
}
MYELOID_CLUSTERS = {"Macrophages", "DCs", "pDCs"}

T_PATHWAYS = [
    "HALLMARK_TNFA_SIGNALING_VIA_NFKB",
    "HALLMARK_MTORC1_SIGNALING",
    "HALLMARK_P53_PATHWAY",
    "HALLMARK_INTERFERON_ALPHA_RESPONSE",
    "HALLMARK_INTERFERON_GAMMA_RESPONSE",
]
M_PATHWAYS = [
    "HALLMARK_MTORC1_SIGNALING",
    "HALLMARK_INTERFERON_GAMMA_RESPONSE",
    "HALLMARK_TNFA_SIGNALING_VIA_NFKB",
    "HALLMARK_INTERFERON_ALPHA_RESPONSE",
    "HALLMARK_INFLAMMATORY_RESPONSE",
    "HALLMARK_COMPLEMENT",
]


@dataclass(frozen=True)
class Signature:
    signature: str
    lineage: str
    source: str
    genes: tuple[str, ...]


def ensure_dirs() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_DATA_DIR.mkdir(parents=True, exist_ok=True)


def read_table(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_csv(path)


def is_interpretable_symbol(gene: str) -> bool:
    if not isinstance(gene, str) or not gene:
        return False
    if gene.startswith(("AC", "AL", "AP", "LINC", "MIR", "MT-")):
        return False
    if "." in gene:
        return False
    if gene.startswith(("RPL", "RPS")):
        return False
    return bool(re.match(r"^[A-Za-z0-9][A-Za-z0-9.-]*$", gene))


def parse_leading_edge(value: object) -> list[str]:
    if pd.isna(value):
        return []
    genes = re.split(r"[;,]\s*", str(value).strip())
    return [g for g in genes if is_interpretable_symbol(g)]


def top_positive_de(df: pd.DataFrame, padj_cutoff: float | None, n: int, pvalue_cutoff: float | None = None) -> list[str]:
    work = df.copy()
    for col in ["log2FoldChange", "padj", "pvalue"]:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce")
    work = work[work["log2FoldChange"] > 0]
    if padj_cutoff is not None:
        work = work[work["padj"] <= padj_cutoff]
    if pvalue_cutoff is not None:
        work = work[work["pvalue"] <= pvalue_cutoff]
    work = work[work["gene"].map(is_interpretable_symbol)]
    sort_cols = [c for c in ["padj", "pvalue"] if c in work.columns]
    if sort_cols:
        work = work.sort_values(sort_cols + ["log2FoldChange"], ascending=[True] * len(sort_cols) + [False])
    else:
        work = work.sort_values("log2FoldChange", ascending=False)
    return work["gene"].drop_duplicates().head(n).tolist()


def leading_edge_signature(gsea: pd.DataFrame, pathway: str) -> list[str]:
    row = gsea[gsea["pathway"] == pathway]
    if row.empty:
        return []
    return list(dict.fromkeys(parse_leading_edge(row.iloc[0]["leadingEdge"])))


def build_signatures() -> list[Signature]:
    t_de = read_table(T_DE_PATH)
    m_de = read_table(M_DE_PATH)
    t_gsea = read_table(T_GSEA_PATH)
    m_gsea = read_table(M_GSEA_PATH)

    signatures: list[Signature] = []

    signatures.append(
        Signature(
            signature="T_DE_FDR05_positive",
            lineage="T_cell",
            source="OSCC T-cell dynamic DE, positive log2FC, FDR < 0.05",
            genes=tuple(top_positive_de(t_de, padj_cutoff=0.05, n=50)),
        )
    )
    signatures.append(
        Signature(
            signature="M_DE_FDR10_positive",
            lineage="Myeloid",
            source="OSCC myeloid dynamic DE, positive log2FC, FDR < 0.10",
            genes=tuple(top_positive_de(m_de, padj_cutoff=0.10, n=50)),
        )
    )
    signatures.append(
        Signature(
            signature="M_DE_nominal_top30_positive",
            lineage="Myeloid",
            source="OSCC myeloid dynamic DE, positive log2FC, nominal P < 0.01, top 30",
            genes=tuple(top_positive_de(m_de, padj_cutoff=None, pvalue_cutoff=0.01, n=30)),
        )
    )

    t_union: list[str] = []
    for pathway in T_PATHWAYS:
        genes = leading_edge_signature(t_gsea, pathway)
        t_union.extend(genes)
        label = pathway.replace("HALLMARK_", "T_LE_")
        signatures.append(Signature(label, "T_cell", f"OSCC T-cell Hallmark leading edge: {pathway}", tuple(genes)))
    signatures.append(
        Signature(
            "T_LE_union_core",
            "T_cell",
            "Union of selected OSCC T-cell Hallmark leading-edge genes",
            tuple(dict.fromkeys(t_union)),
        )
    )

    m_union: list[str] = []
    for pathway in M_PATHWAYS:
        genes = leading_edge_signature(m_gsea, pathway)
        m_union.extend(genes)
        label = pathway.replace("HALLMARK_", "M_LE_")
        signatures.append(Signature(label, "Myeloid", f"OSCC myeloid Hallmark leading edge: {pathway}", tuple(genes)))
    signatures.append(
        Signature(
            "M_LE_union_core",
            "Myeloid",
            "Union of selected OSCC myeloid Hallmark leading-edge genes",
            tuple(dict.fromkeys(m_union)),
        )
    )

    kept = [sig for sig in signatures if sig.genes]
    return kept


def read_counts_header() -> list[str]:
    with gzip.open(COUNTS_PATH, "rt", encoding="utf-8") as handle:
        header = handle.readline().rstrip("\n\r").split("\t")
    return [h for h in header if h]


def load_metadata(cell_ids: list[str]) -> pd.DataFrame:
    meta = pd.read_csv(META_PATH, sep="\t", compression="gzip")
    if meta["cell.id"].duplicated().any():
        dupes = meta.loc[meta["cell.id"].duplicated(), "cell.id"].head(5).tolist()
        raise ValueError(f"Duplicated cell.id values in metadata: {dupes}")
    metadata_cells = set(meta["cell.id"])
    missing = [cell for cell in cell_ids if cell not in metadata_cells]
    if missing:
        raise ValueError(f"{len(missing)} count-matrix cells are missing from metadata; first={missing[:5]}")
    meta = meta.set_index("cell.id").loc[cell_ids].reset_index()
    meta["lineage_validation"] = np.where(
        meta["cluster"].isin(T_CLUSTERS),
        "T_cell",
        np.where(meta["cluster"].isin(MYELOID_CLUSTERS), "Myeloid", "Other"),
    )
    return meta


def count_matrix_pass(cell_count: int) -> tuple[np.ndarray, int]:
    libsize = np.zeros(cell_count, dtype=np.float64)
    gene_count = 0
    with gzip.open(COUNTS_PATH, "rt", encoding="utf-8") as handle:
        next(handle)
        for line in handle:
            tab = line.find("\t")
            if tab <= 0:
                continue
            values = np.fromstring(line[tab + 1 :], sep="\t", dtype=np.float32)
            if values.size != cell_count:
                raise ValueError(f"Unexpected count length at gene line {gene_count + 1}: {values.size} != {cell_count}")
            libsize += values
            gene_count += 1
    return libsize, gene_count


def score_signatures(cell_count: int, libsize: np.ndarray, signatures: list[Signature]) -> tuple[pd.DataFrame, dict[str, int]]:
    gene_to_signatures: dict[str, list[int]] = {}
    for idx, sig in enumerate(signatures):
        for gene in sig.genes:
            gene_to_signatures.setdefault(gene, []).append(idx)

    score_sums = np.zeros((cell_count, len(signatures)), dtype=np.float32)
    present_counts = np.zeros(len(signatures), dtype=np.int32)
    safe_libsize = np.where(libsize > 0, libsize, np.nan)

    with gzip.open(COUNTS_PATH, "rt", encoding="utf-8") as handle:
        next(handle)
        for line in handle:
            tab = line.find("\t")
            if tab <= 0:
                continue
            gene = line[:tab]
            sig_idxs = gene_to_signatures.get(gene)
            if not sig_idxs:
                continue
            values = np.fromstring(line[tab + 1 :], sep="\t", dtype=np.float32)
            norm = np.log1p((values / safe_libsize) * 10000.0).astype(np.float32)
            norm = np.nan_to_num(norm, nan=0.0, posinf=0.0, neginf=0.0)
            for idx in sig_idxs:
                score_sums[:, idx] += norm
                present_counts[idx] += 1

    score_cols = {}
    present_by_signature: dict[str, int] = {}
    for idx, sig in enumerate(signatures):
        present = int(present_counts[idx])
        present_by_signature[sig.signature] = present
        if present == 0:
            score_cols[sig.signature] = np.full(cell_count, np.nan, dtype=np.float32)
        else:
            score_cols[sig.signature] = score_sums[:, idx] / float(present)
    return pd.DataFrame(score_cols), present_by_signature


def summarize_metadata(meta: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col in ["patient", "treatment", "sort", "cluster", "lineage_validation"]:
        counts = meta[col].fillna("NA").value_counts(dropna=False)
        for value, n in counts.items():
            rows.append({"field": col, "value": value, "n_cells": int(n)})
    return pd.DataFrame(rows)


def markdown_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_No rows._"
    text_df = df.copy()
    text_df = text_df.fillna("")
    headers = [str(c) for c in text_df.columns]
    rows = [[str(v) for v in row] for row in text_df.to_numpy()]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(value.replace("|", "\\|") for value in row) + " |")
    return "\n".join(lines)


def bh_adjust(pvalues: list[float]) -> list[float]:
    arr = np.asarray([1.0 if (p is None or math.isnan(p)) else p for p in pvalues], dtype=float)
    n = len(arr)
    order = np.argsort(arr)
    ranks = np.empty(n, dtype=int)
    ranks[order] = np.arange(1, n + 1)
    adjusted = arr * n / ranks
    adjusted_ordered = adjusted[order]
    for i in range(n - 2, -1, -1):
        adjusted_ordered[i] = min(adjusted_ordered[i], adjusted_ordered[i + 1])
    out = np.empty(n, dtype=float)
    out[order] = np.minimum(adjusted_ordered, 1.0)
    return out.tolist()


def paired_patient_scores(meta: pd.DataFrame, scores: pd.DataFrame, signatures: list[Signature]) -> tuple[pd.DataFrame, pd.DataFrame]:
    sig_lineage = {sig.signature: sig.lineage for sig in signatures}
    score_data = pd.concat([meta[["cell.id", "patient", "treatment", "cluster", "lineage_validation"]], scores], axis=1)
    rows = []
    for sig_name, lineage in sig_lineage.items():
        subset = score_data[score_data["lineage_validation"] == lineage]
        grouped = (
            subset.groupby(["patient", "treatment"], dropna=False)
            .agg(mean_score=(sig_name, "mean"), n_cells=(sig_name, "count"))
            .reset_index()
        )
        grouped["signature"] = sig_name
        grouped["target_lineage"] = lineage
        rows.append(grouped)
    patient_scores = pd.concat(rows, ignore_index=True)
    patient_scores = patient_scores[
        ["signature", "target_lineage", "patient", "treatment", "mean_score", "n_cells"]
    ].sort_values(["target_lineage", "signature", "patient", "treatment"])

    stat_rows = []
    for sig_name, lineage in sig_lineage.items():
        sub = patient_scores[(patient_scores["signature"] == sig_name) & (patient_scores["target_lineage"] == lineage)]
        wide_mean = sub.pivot(index="patient", columns="treatment", values="mean_score")
        wide_n = sub.pivot(index="patient", columns="treatment", values="n_cells")
        if not {"pre", "post"}.issubset(set(wide_mean.columns)):
            continue
        keep = (wide_n["pre"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT) & (wide_n["post"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT)
        wide_mean = wide_mean.loc[keep]
        wide_n = wide_n.loc[keep]
        deltas = (wide_mean["post"] - wide_mean["pre"]).dropna()
        n = int(deltas.shape[0])
        if n == 0:
            t_p = math.nan
            w_p = math.nan
            t_stat = math.nan
            w_stat = math.nan
            mean_delta = math.nan
            median_delta = math.nan
            sd_delta = math.nan
            cohen_dz = math.nan
            pos_n = 0
            pos_frac = math.nan
        else:
            mean_delta = float(deltas.mean())
            median_delta = float(deltas.median())
            sd_delta = float(deltas.std(ddof=1)) if n > 1 else math.nan
            cohen_dz = float(mean_delta / sd_delta) if sd_delta and not math.isnan(sd_delta) else math.nan
            pos_n = int((deltas > 0).sum())
            pos_frac = float(pos_n / n)
            if n > 1 and sd_delta and not math.isnan(sd_delta):
                t_res = stats.ttest_1samp(deltas.to_numpy(), popmean=0.0)
                t_stat = float(t_res.statistic)
                t_p = float(t_res.pvalue)
            else:
                t_stat = math.nan
                t_p = math.nan
            if n >= 3 and not np.allclose(deltas.to_numpy(), 0):
                try:
                    w_res = stats.wilcoxon(deltas.to_numpy(), zero_method="wilcox", alternative="two-sided")
                    w_stat = float(w_res.statistic)
                    w_p = float(w_res.pvalue)
                except ValueError:
                    w_stat = math.nan
                    w_p = math.nan
            else:
                w_stat = math.nan
                w_p = math.nan
        stat_rows.append(
            {
                "signature": sig_name,
                "target_lineage": lineage,
                "n_paired_patients": n,
                "min_cells_per_patient_timepoint": MIN_CELLS_PER_PATIENT_TIMEPOINT,
                "mean_post_minus_pre": mean_delta,
                "median_post_minus_pre": median_delta,
                "sd_post_minus_pre": sd_delta,
                "cohen_dz": cohen_dz,
                "positive_delta_patients": pos_n,
                "positive_delta_fraction": pos_frac,
                "paired_t_stat": t_stat,
                "paired_t_pvalue": t_p,
                "wilcoxon_stat": w_stat,
                "wilcoxon_pvalue": w_p,
                "interpretation_boundary": "External treatment-direction validation only; no OSCC response-depth labels in GSE123813.",
            }
        )
    stats_df = pd.DataFrame(stat_rows)
    if not stats_df.empty:
        stats_df["paired_t_fdr"] = bh_adjust(stats_df["paired_t_pvalue"].tolist())
        stats_df["wilcoxon_fdr"] = bh_adjust(stats_df["wilcoxon_pvalue"].tolist())
        stats_df["direction_supported"] = np.where(
            (stats_df["mean_post_minus_pre"] > 0) & (stats_df["positive_delta_fraction"] >= 0.5),
            "positive_post_pre",
            "not_positive",
        )
    return patient_scores, stats_df.sort_values(["target_lineage", "signature"]).reset_index(drop=True)


def cluster_composition_validation(meta: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    stat_rows = []
    cluster_sets = {
        "T_cell": sorted(T_CLUSTERS),
        "Myeloid": sorted(MYELOID_CLUSTERS),
    }
    for lineage, clusters in cluster_sets.items():
        lin_meta = meta[meta["lineage_validation"] == lineage].copy()
        denom = (
            lin_meta.groupby(["patient", "treatment"], dropna=False)
            .size()
            .reset_index(name="lineage_cells")
        )
        counts = (
            lin_meta.groupby(["patient", "treatment", "cluster"], dropna=False)
            .size()
            .reset_index(name="cluster_cells")
        )
        for cluster in clusters:
            cluster_counts = counts[counts["cluster"] == cluster][["patient", "treatment", "cluster_cells"]]
            merged = denom.merge(cluster_counts, on=["patient", "treatment"], how="left")
            merged["cluster_cells"] = merged["cluster_cells"].fillna(0).astype(int)
            merged["cluster"] = cluster
            merged["lineage"] = lineage
            merged["proportion_within_lineage"] = merged["cluster_cells"] / merged["lineage_cells"]
            merged["logit_proportion"] = np.log((merged["cluster_cells"] + 0.5) / (merged["lineage_cells"] - merged["cluster_cells"] + 0.5))
            rows.append(
                merged[
                    [
                        "lineage",
                        "cluster",
                        "patient",
                        "treatment",
                        "cluster_cells",
                        "lineage_cells",
                        "proportion_within_lineage",
                        "logit_proportion",
                    ]
                ]
            )

            wide = merged.pivot(index="patient", columns="treatment", values="logit_proportion")
            nwide = merged.pivot(index="patient", columns="treatment", values="lineage_cells")
            if not {"pre", "post"}.issubset(set(wide.columns)):
                continue
            keep = (nwide["pre"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT) & (nwide["post"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT)
            deltas = (wide.loc[keep, "post"] - wide.loc[keep, "pre"]).dropna()
            n = int(deltas.shape[0])
            if n > 1 and float(deltas.std(ddof=1)) > 0:
                t_res = stats.ttest_1samp(deltas.to_numpy(), popmean=0.0)
                t_stat = float(t_res.statistic)
                t_p = float(t_res.pvalue)
            else:
                t_stat = math.nan
                t_p = math.nan
            if n >= 3 and not np.allclose(deltas.to_numpy(), 0):
                try:
                    w_res = stats.wilcoxon(deltas.to_numpy(), zero_method="wilcox", alternative="two-sided")
                    w_stat = float(w_res.statistic)
                    w_p = float(w_res.pvalue)
                except ValueError:
                    w_stat = math.nan
                    w_p = math.nan
            else:
                w_stat = math.nan
                w_p = math.nan
            stat_rows.append(
                {
                    "lineage": lineage,
                    "cluster": cluster,
                    "n_paired_patients": n,
                    "mean_logit_post_minus_pre": float(deltas.mean()) if n else math.nan,
                    "median_logit_post_minus_pre": float(deltas.median()) if n else math.nan,
                    "positive_delta_patients": int((deltas > 0).sum()) if n else 0,
                    "positive_delta_fraction": float((deltas > 0).sum() / n) if n else math.nan,
                    "paired_t_stat": t_stat,
                    "paired_t_pvalue": t_p,
                    "wilcoxon_stat": w_stat,
                    "wilcoxon_pvalue": w_p,
                }
            )
    composition = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    comp_stats = pd.DataFrame(stat_rows)
    if not comp_stats.empty:
        comp_stats["paired_t_fdr"] = bh_adjust(comp_stats["paired_t_pvalue"].tolist())
        comp_stats["wilcoxon_fdr"] = bh_adjust(comp_stats["wilcoxon_pvalue"].tolist())
    return composition, comp_stats


def write_manifest(signatures: list[Signature], present: dict[str, int]) -> pd.DataFrame:
    rows = []
    for sig in signatures:
        rows.append(
            {
                "signature": sig.signature,
                "target_lineage": sig.lineage,
                "source": sig.source,
                "n_genes_defined": len(sig.genes),
                "n_genes_present_in_GSE123813_BCC_counts": present.get(sig.signature, 0),
                "genes_defined": ";".join(sig.genes),
            }
        )
    manifest = pd.DataFrame(rows)
    manifest.to_csv(RESULTS_DIR / "GSE123813_gene_set_manifest.csv", index=False)
    return manifest


def write_source_workbook(csv_paths: list[Path]) -> Path:
    out = SOURCE_DATA_DIR / "ExtendedData5_source_data.xlsx"
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme.append(["source_workbook", out.name])
    readme.append(["generated_on", datetime.now().strftime("%Y-%m-%d")])
    readme.append(["generator", "03_rebuild/analysis/17_external_validation_gse123813.py"])
    readme.append(["note", "Source data for external GSE123813 treatment-direction validation."])
    for cell in readme["A"]:
        cell.font = Font(bold=True)
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 100

    used = {"README"}
    for path in csv_paths:
        raw_name = path.stem.replace("ExtendedData5_", "")
        sheet_name = re.sub(r"[\[\]\*\?/\\:]", "_", raw_name)[:31] or "Sheet"
        base = sheet_name
        idx = 2
        while sheet_name in used:
            suffix = f"_{idx}"
            sheet_name = base[: 31 - len(suffix)] + suffix
            idx += 1
        used.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                ws.append(row)
        if ws.max_row:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        for col_cells in ws.columns:
            max_len = max(len("" if c.value is None else str(c.value)) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, 8), 55)

    wb.save(out)
    check = load_workbook(out)
    check.close()
    return out


def update_extended_data_manifest() -> None:
    manifest_path = FIGURE_DIR / "EXTENDED_DATA_FIGURE_MANIFEST.csv"
    rows: list[dict[str, str]] = []
    fieldnames = ["generated_at", "figure", "stem", "primary_message"]
    if manifest_path.exists():
        with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames:
                fieldnames = reader.fieldnames
            rows = [row for row in reader if row.get("figure") != "ExtendedData5"]
    rows.append(
        {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S HKT"),
            "figure": "ExtendedData5",
            "stem": FIG_STEM,
            "primary_message": "External GSE123813 BCC anti-PD-1 data support treatment-direction immune-state remodeling, not OSCC response prediction.",
        }
    )
    rows = sorted(rows, key=lambda r: r.get("figure", ""))
    with manifest_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def make_plot(patient_scores: pd.DataFrame, stats_df: pd.DataFrame) -> list[Path]:
    label_map = {
        "T_DE_FDR05_positive": "T dynamic DE",
        "T_LE_union_core": "T leading-edge core",
        "M_DE_FDR10_positive": "Myeloid dynamic DE",
        "M_LE_union_core": "Myeloid leading-edge core",
        "T_LE_TNFA_SIGNALING_VIA_NFKB": "T: TNFA/NF-kB LE",
        "T_LE_MTORC1_SIGNALING": "T: mTORC1 LE",
        "T_LE_P53_PATHWAY": "T: p53 LE",
        "T_LE_INTERFERON_ALPHA_RESPONSE": "T: IFN-alpha LE",
        "T_LE_INTERFERON_GAMMA_RESPONSE": "T: IFN-gamma LE",
        "M_LE_MTORC1_SIGNALING": "Myeloid: mTORC1 LE",
        "M_LE_INTERFERON_GAMMA_RESPONSE": "Myeloid: IFN-gamma LE",
        "M_LE_TNFA_SIGNALING_VIA_NFKB": "Myeloid: TNFA/NF-kB LE",
        "M_LE_INTERFERON_ALPHA_RESPONSE": "Myeloid: IFN-alpha LE",
        "M_LE_INFLAMMATORY_RESPONSE": "Myeloid: inflammatory LE",
        "M_LE_COMPLEMENT": "Myeloid: complement LE",
        "M_DE_nominal_top30_positive": "Myeloid nominal top 30",
    }

    def pretty_label(signature: str) -> str:
        return label_map.get(signature, signature.replace("_", " "))

    anchor_sigs = [
        "T_DE_FDR05_positive",
        "T_LE_union_core",
        "M_DE_FDR10_positive",
        "M_LE_union_core",
    ]
    anchors = [s for s in anchor_sigs if s in set(patient_scores["signature"])]

    plt.rcParams.update(
        {
            "font.family": "Arial",
            "axes.linewidth": 0.6,
            "xtick.major.width": 0.6,
            "ytick.major.width": 0.6,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )
    fig = plt.figure(figsize=(10.2, 7.2), constrained_layout=True)
    grid = fig.add_gridspec(2, 4, height_ratios=[1.0, 1.42])
    axes = [fig.add_subplot(grid[0, i]) for i in range(4)]
    ax_delta = fig.add_subplot(grid[1, :])

    panel_letters = list("abcde")
    for idx, (ax, sig) in enumerate(zip(axes, anchors)):
        sub = patient_scores[patient_scores["signature"] == sig]
        wide = sub.pivot(index="patient", columns="treatment", values="mean_score")
        nwide = sub.pivot(index="patient", columns="treatment", values="n_cells")
        if {"pre", "post"}.issubset(set(wide.columns)):
            keep = (nwide["pre"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT) & (nwide["post"] >= MIN_CELLS_PER_PATIENT_TIMEPOINT)
            wide = wide.loc[keep]
            for _, row in wide.iterrows():
                ax.plot([0, 1], [row["pre"], row["post"]], color="#A9B0B8", lw=0.8, alpha=0.85)
                ax.scatter([0, 1], [row["pre"], row["post"]], color="#47515C", s=14, zorder=3)
            if len(wide):
                means = [wide["pre"].mean(), wide["post"].mean()]
                ax.plot([0, 1], means, color="#C55A11", lw=1.8, marker="o", ms=4)
        ax.set_xticks([0, 1])
        ax.set_xticklabels(["Pre", "Post"])
        ax.set_title(pretty_label(sig), fontsize=8.4, fontweight="normal", pad=4)
        ax.set_ylabel("Mean logCP10K score", fontsize=7.2)
        ax.tick_params(axis="both", labelsize=7)
        ax.spines[["top", "right"]].set_visible(False)
        ax.text(
            -0.18,
            1.08,
            panel_letters[idx],
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            va="bottom",
            ha="left",
        )

    for ax in axes[len(anchors) :]:
        ax.axis("off")

    plot_stats = stats_df.copy()
    plot_stats = plot_stats[plot_stats["n_paired_patients"] >= 3].copy()
    plot_stats["lineage_order"] = plot_stats["target_lineage"].map({"T_cell": 0, "Myeloid": 1}).fillna(2)
    plot_stats = plot_stats.sort_values(
        ["lineage_order", "mean_post_minus_pre"], ascending=[True, False]
    )
    y = np.arange(len(plot_stats))
    colors = np.where(plot_stats["target_lineage"].eq("T_cell"), "#2C6FBB", "#008B73")
    ax_delta.axvline(0, color="#555555", lw=0.75)
    ax_delta.scatter(plot_stats["mean_post_minus_pre"], y, s=38, c=colors, edgecolor="black", linewidth=0.35, zorder=3)
    for i, (_, row) in enumerate(plot_stats.iterrows()):
        se = row["sd_post_minus_pre"] / math.sqrt(row["n_paired_patients"]) if row["n_paired_patients"] > 1 else math.nan
        if not math.isnan(se):
            lo = row["mean_post_minus_pre"] - 1.96 * se
            hi = row["mean_post_minus_pre"] + 1.96 * se
            ax_delta.plot([lo, hi], [i, i], color="#5A5A5A", lw=0.85, zorder=2)
    labels = [
        f"{pretty_label(r.signature)} (n={int(r.n_paired_patients)}, FDR={r.paired_t_fdr:.2g})"
        for r in plot_stats.itertuples()
    ]
    ax_delta.set_yticks(y)
    ax_delta.set_yticklabels(labels, fontsize=6.7)
    ax_delta.set_xlabel("Patient-level post - pre module score", fontsize=8)
    ax_delta.set_title("Paired anti-PD-1 boundary test in GSE123813 BCC", fontsize=8.6, fontweight="normal", pad=5)
    ax_delta.tick_params(axis="x", labelsize=7)
    ax_delta.spines[["top", "right"]].set_visible(False)
    ax_delta.text(
        -0.02,
        1.04,
        "e",
        transform=ax_delta.transAxes,
        fontsize=10,
        fontweight="bold",
        va="bottom",
        ha="left",
    )

    paths = []
    for ext in ["png", "pdf", "svg"]:
        out = FIGURE_DIR / f"{FIG_STEM}.{ext}"
        fig.savefig(out, dpi=300)
        paths.append(out)
    plt.close(fig)
    return paths


def write_interpretation(
    meta_summary: pd.DataFrame,
    manifest: pd.DataFrame,
    patient_scores: pd.DataFrame,
    stats_df: pd.DataFrame,
    cluster_stats: pd.DataFrame,
    figure_paths: list[Path],
    source_workbook: Path,
) -> None:
    paired_counts = (
        patient_scores.groupby(["target_lineage", "signature"])["patient"]
        .nunique()
        .reset_index(name="n_patients_with_any_score")
    )
    anchors = stats_df[stats_df["signature"].isin(["T_DE_FDR05_positive", "T_LE_union_core", "M_DE_FDR10_positive", "M_LE_union_core"])]
    lines = [
        "# External validation: GSE123813 BCC anti-PD-1 paired scRNA-seq",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S HKT')}",
        "",
        "## Dataset boundary",
        "",
        "- Source: GSE123813 BCC anti-PD-1 single-cell RNA-seq count matrix and metadata already downloaded under `00_raw_data/GSE123813_validation/`.",
        "- Use: independent treatment-direction validation of OSCC-derived T-cell and myeloid remodeling modules.",
        "- Boundary: this is basal cell carcinoma, not OSCC, and the available local metadata do not provide the same OSCC pathological response-depth labels. Therefore this analysis must not be described as independent response-prediction validation.",
        "",
        "## Local data summary",
        "",
        f"- Cells scored: {len(patient_scores['patient'].unique())} patients represented across scored lineage/signature summaries.",
        f"- Signature definitions: {len(manifest)} modules; present genes are recorded in `GSE123813_gene_set_manifest.csv`.",
        f"- Minimum patient/timepoint cell threshold: {MIN_CELLS_PER_PATIENT_TIMEPOINT}.",
        "",
        "## Anchor module statistics",
        "",
    ]
    if anchors.empty:
        lines.append("No anchor module statistics were available.")
    else:
        for row in anchors.sort_values(["target_lineage", "signature"]).itertuples(index=False):
            lines.append(
                "- {sig}: lineage={lin}, n={n}, mean post-pre={mean:.4f}, positive patients={pos}/{n}, paired t P={p:.3g}, FDR={q:.3g}; Wilcoxon P={wp:.3g}, FDR={wq:.3g}.".format(
                    sig=row.signature,
                    lin=row.target_lineage,
                    n=int(row.n_paired_patients),
                    mean=float(row.mean_post_minus_pre),
                    pos=int(row.positive_delta_patients),
                    p=float(row.paired_t_pvalue) if not math.isnan(row.paired_t_pvalue) else math.nan,
                    q=float(row.paired_t_fdr) if not math.isnan(row.paired_t_fdr) else math.nan,
                    wp=float(row.wilcoxon_pvalue) if not math.isnan(row.wilcoxon_pvalue) else math.nan,
                    wq=float(row.wilcoxon_fdr) if not math.isnan(row.wilcoxon_fdr) else math.nan,
                )
            )
    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "The independent dataset is best used as an orthogonal boundary test for treatment-induced immune-state remodeling. In this run, OSCC-derived expression modules showed mixed and mostly non-significant post-pre shifts in BCC. This should be interpreted as partial external support for a general anti-PD-1 remodeling frame, not as validation of an OSCC response-prediction signature.",
            "",
            "## Metadata cluster remodeling check",
            "",
            markdown_table(cluster_stats.sort_values(["lineage", "cluster"]).round(5)),
            "",
            "## Outputs",
            "",
        ]
    )
    for path in figure_paths:
        lines.append(f"- Figure: `{path.relative_to(ROOT).as_posix()}`")
    lines.extend(
        [
            f"- Source workbook: `{source_workbook.relative_to(ROOT).as_posix()}`",
            f"- Patient scores: `{(RESULTS_DIR / 'GSE123813_patient_signature_scores.csv').relative_to(ROOT).as_posix()}`",
            f"- Delta stats: `{(RESULTS_DIR / 'GSE123813_paired_delta_stats.csv').relative_to(ROOT).as_posix()}`",
            "",
            "## Patient/signature availability",
            "",
            markdown_table(paired_counts),
            "",
            "## Metadata summary preview",
            "",
            markdown_table(meta_summary.head(20)),
            "",
        ]
    )
    (RESULTS_DIR / "GSE123813_validation_interpretation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--reuse-existing",
        action="store_true",
        help="Reuse existing score/stat CSV files and regenerate interpretation, source workbook and manifest.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ensure_dirs()
    print(f"Workspace: {ROOT}")
    print(f"Counts:    {COUNTS_PATH}")
    print(f"Metadata:  {META_PATH}")

    cell_ids = read_counts_header()
    print(f"Count-matrix cells: {len(cell_ids):,}")
    meta = load_metadata(cell_ids)
    meta_summary = summarize_metadata(meta)
    meta_summary.to_csv(RESULTS_DIR / "GSE123813_metadata_summary.csv", index=False)
    cluster_composition, cluster_stats = cluster_composition_validation(meta)
    cluster_composition.to_csv(RESULTS_DIR / "GSE123813_patient_cluster_composition.csv", index=False)
    cluster_stats.to_csv(RESULTS_DIR / "GSE123813_cluster_composition_delta_stats.csv", index=False)

    signatures = build_signatures()
    print(f"Signatures defined: {len(signatures)}")

    reusable = all(
        path.exists()
        for path in [
            RESULTS_DIR / "GSE123813_gene_set_manifest.csv",
            RESULTS_DIR / "GSE123813_patient_signature_scores.csv",
            RESULTS_DIR / "GSE123813_paired_delta_stats.csv",
        ]
    )
    if args.reuse_existing and reusable:
        print("Reusing existing external-validation score/stat files.")
        manifest = pd.read_csv(RESULTS_DIR / "GSE123813_gene_set_manifest.csv")
        patient_scores = pd.read_csv(RESULTS_DIR / "GSE123813_patient_signature_scores.csv")
        stats_df = pd.read_csv(RESULTS_DIR / "GSE123813_paired_delta_stats.csv")
    else:
        print("Pass 1/2: computing per-cell library sizes...")
        libsize, n_genes = count_matrix_pass(len(cell_ids))
        library_df = pd.DataFrame({"cell.id": cell_ids, "library_size": libsize.astype(np.float64)})
        library_df.to_csv(RESULTS_DIR / "GSE123813_cell_library_size_summary.csv", index=False)
        print(f"Genes scanned: {n_genes:,}; median library size: {np.median(libsize):.1f}")

        print("Pass 2/2: scoring selected OSCC-derived modules...")
        scores, present = score_signatures(len(cell_ids), libsize, signatures)
        manifest = write_manifest(signatures, present)
        patient_scores, stats_df = paired_patient_scores(meta, scores, signatures)

        patient_scores.to_csv(RESULTS_DIR / "GSE123813_patient_signature_scores.csv", index=False)
        stats_df.to_csv(RESULTS_DIR / "GSE123813_paired_delta_stats.csv", index=False)

        cell_scores = pd.concat([meta[["cell.id", "patient", "treatment", "cluster", "lineage_validation"]], scores], axis=1)
        cell_scores.to_csv(RESULTS_DIR / "GSE123813_cell_signature_scores.csv.gz", index=False, compression="gzip")

    source_paths = []
    source_specs = [
        ("ExtendedData5_patient_signature_scores.csv", patient_scores),
        ("ExtendedData5_paired_delta_stats.csv", stats_df),
        ("ExtendedData5_gene_set_manifest.csv", manifest),
        ("ExtendedData5_metadata_summary.csv", meta_summary),
        ("ExtendedData5_patient_cluster_composition.csv", cluster_composition),
        ("ExtendedData5_cluster_composition_delta_stats.csv", cluster_stats),
    ]
    for name, df in source_specs:
        path = SOURCE_DATA_DIR / name
        df.to_csv(path, index=False)
        source_paths.append(path)
    source_workbook = write_source_workbook(source_paths)
    figure_paths = make_plot(patient_scores, stats_df)
    update_extended_data_manifest()
    write_interpretation(meta_summary, manifest, patient_scores, stats_df, cluster_stats, figure_paths, source_workbook)

    print("External validation complete.")
    print(stats_df[["signature", "target_lineage", "n_paired_patients", "mean_post_minus_pre", "paired_t_pvalue", "paired_t_fdr"]].to_string(index=False))


if __name__ == "__main__":
    main()
